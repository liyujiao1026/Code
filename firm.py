import openpyxl
import numpy
import nltk
import Levenshtein as l

## 1.Reading the test.xlsx file
datapath="/Users/Yujiao/Desktop/test.xlsx"
work_book = openpyxl.load_workbook(datapath)
sheet_name=work_book.get_active_sheet().title

work_sheet = work_book.get_sheet_by_name(sheet_name)
print ('Sheet_%s is the current working sheet'%sheet_name)

testcell=work_sheet.cell('B3').value
print('%s is the the content of B3 cell'%testcell)

##=====================================================================##
## 2.Extract the column of business description

for r in xrange(1,4):
#    for c in [1,47]:
    val=work_sheet.cell(row=r,column=47).value
    if val != None:
        print '%d:  %s\n'%(r,val)



##======================================================================##
## 3. Corpus
class MyCorpus(object):
    def __iter__(self):
        for r in xrange(0,20):
            yield work_sheet.cell(row=r,column=47).value


from gensim import corpora, models, similarities

Corp=MyCorpus()
corpusIte=iter(Corp)
a=corpusIte.next()
b=corpusIte.next()
c=corpusIte.next()
print l.ratio(b,c)
print l.distance(b,c)


#
def func():
	"ehte"
	list(None)







            
